import openpyxl

path = "D:\certificates\Selenium\data1.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active #workbook.get_sheet_by_name("Sheet name")

rows = sheet.max_row
cols = sheet.max_column

print("Rows:{0}, Columns:{1}".format(rows, cols))


for c in range(1,cols+1):
    print(sheet.cell(row=1,column=c).value,end="   ")
print()

for r in range(2,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="        ")
    print()
